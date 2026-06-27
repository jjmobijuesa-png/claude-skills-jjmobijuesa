# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "E:/vars/var 8/3 Analisis Financiero QVP/ANALISIS_FINANCIERO_QUEVEPALMA_2024_2025.xlsx"
NAVY="1B2A4A"; STEEL="4A6FA5"; GOLD="C5A55A"; GREEN="1B7A43"; RED="B02418"; LIGHT="EEF2F7"
fill_navy=PatternFill("solid",fgColor=NAVY); fill_steel=PatternFill("solid",fgColor=STEEL)
fill_gold=PatternFill("solid",fgColor=GOLD); fill_light=PatternFill("solid",fgColor=LIGHT)
f_sub=Font(color="FFFFFF",bold=True,size=10)
thin=Side(style="thin",color="B0B8C4"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center",vertical="center"); wrap=Alignment(wrap_text=True,vertical="top")
NUM="#,##0.00"; PCT="0.00%"

wb=load_workbook(OUT)
ws=wb.create_sheet("Forense Oro Rojo")
ws.merge_cells("A1:D1"); c=ws["A1"]; c.value="ANALISIS FORENSE - PRESTAMO RELACIONADO ORO ROJO / OROJUEZ S.A."
c.font=Font(color="FFFFFF",bold=True,size=13); c.fill=fill_navy; c.alignment=ctr
ws.merge_cells("A2:D2"); c=ws["A2"]; c.value="Cuenta contable 1240101012 - Companias Relacionadas (Econ. David Juez)  |  Computos verificados con Wolfram"
c.font=Font(color="FFFFFF",bold=True,size=9); c.fill=fill_steel; c.alignment=ctr

rows=[
 ("ORIGEN Y ESTRUCTURA DEL PRESTAMO","sec",None,None),
 ("Capital prestado (15-may-2023)",1551058.82,NUM,""),
 ("Interes/credito unico (18-may-2023)",150000.00,NUM,"una sola vez"),
 ("Total deuda registrada",1701058.82,NUM,"tot"),
 ("Cuota mensual fija",28350.98,NUM,""),
 ("Plazo (cuotas)",60,"0","= 5 anios"),
 ("Maduracion",None,"txt","2028-05-15"),
 ("ESTADO ACTUAL (al 7-may-2026)","sec",None,None),
 ("Total pagado a la fecha",1020635.28,NUM,"42 pagos"),
 ("Saldo pendiente",680423.54,NUM,"tot"),
 ("% del prestamo pagado",0.60,PCT,""),
 ("Cuotas restantes",24,"0","~2 anios mas"),
 ("HALLAZGOS FORENSES (Wolfram)","sec",None,None),
 ("TIR anual efectiva del prestamo",0.0376,PCT,"alerta"),
 ("Costo de deuda de Quevepalma (ref.)",0.10,PCT,"bancos CFN/Pacifico"),
 ("SPREAD = subsidio implicito",0.0624,PCT,"alerta"),
 ("Saldo promedio sobre la vida",850529.41,NUM,""),
 ("COSTO DE OPORTUNIDAD (5 anios)",265512.68,NUM,"alerta"),
 ("Erosion inflacionaria principal (2%/a)",91460.97,NUM,""),
 ("Linea 'DEUDA DE EPACEM' incluida",226831.60,NUM,"cruce caso EPACEM"),
]
r=4
for label,val,fmt,nota in rows:
    if val=="sec":
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
        c=ws.cell(r,1,label); c.font=f_sub; c.fill=fill_steel
    else:
        alert = (nota=="alerta"); tot=(nota=="tot")
        ws.cell(r,1,label).font=Font(bold=tot or alert)
        if fmt=="txt":
            cc=ws.cell(r,2,nota); cc.font=Font(bold=True)
        else:
            cc=ws.cell(r,2,val); cc.number_format=fmt
            if alert: cc.font=Font(bold=True,color=RED)
            elif tot: cc.font=Font(bold=True)
        if nota not in("alerta","tot","txt",""):
            ws.cell(r,3,nota).font=Font(italic=True,size=9,color="777777")
        if alert:
            for col in range(1,4): ws.cell(r,col).fill=PatternFill("solid",fgColor="FBE3E1")
        elif tot:
            for col in range(1,4): ws.cell(r,col).fill=fill_light
    r+=1

r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r+7,end_column=4)
concl=("CONCLUSION FORENSE: Quevepalma mantiene un prestamo a la relacionada Oro Rojo/OroJuez S.A. (Econ. David Juez) "
 "de USD 1,551,058.82 mas un unico cargo de interes de USD 150,000. La TIR efectiva resultante es 3.76% anual, frente a un costo "
 "de deuda de Quevepalma de ~10% (bancos CFN, Banco del Pacifico, Produbanco). El diferencial de 6.24 puntos representa un SUBSIDIO "
 "IMPLICITO a la parte relacionada, con un costo de oportunidad de USD 265,513 sobre la vida del prestamo. "
 "En el contexto de la PERDIDA de USD 1,307,626 de Quevepalma en 2025, este subsidio equivale a ~20% de la perdida anual. "
 "Senales adicionales: (1) existen dos versiones del calculo (FINAL y CONSOLIDADA) con componentes distintos pero identico neto; "
 "(2) la linea 'DEUDA DE EPACEM' (USD 226,831.60) vincula este expediente con el caso forense EPACEM; "
 "(3) no se cobra interes recurrente sobre el saldo decreciente. "
 "NOTA METODOLOGICA: la Ley de Benford NO aplica a este mayor (cuotas identicas y estructuradas); seria aplicable a las facturas "
 "de maquila EPACEM, cuyo archivo requiere descarga adicional.")
c=ws.cell(r,1,concl); c.alignment=wrap; c.font=Font(size=9,color="333333")

ws.column_dimensions["A"].width=44
for col in "BCD": ws.column_dimensions[col].width=18
for row in ws.iter_rows(min_row=3):
    for cc in row:
        if cc.value is not None and cc.row>=3 and cc.row<r: cc.border=border

wb.save(OUT)
print("OK -> pestania 'Forense Oro Rojo' anadida")
print("Pestanas finales:", [s.title for s in wb.worksheets])
