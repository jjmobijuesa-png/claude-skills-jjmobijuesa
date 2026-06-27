# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "E:/vars/var 8/3 Analisis Financiero QVP/ANALISIS_FINANCIERO_QUEVEPALMA_2024_2025.xlsx"

NAVY="1B2A4A"; STEEL="4A6FA5"; GOLD="C5A55A"; GREEN="1B7A43"; RED="B02418"; LIGHT="EEF2F7"
f_sub=Font(color="FFFFFF",bold=True,size=10); f_reg=Font(size=10)
fill_navy=PatternFill("solid",fgColor=NAVY); fill_steel=PatternFill("solid",fgColor=STEEL)
fill_gold=PatternFill("solid",fgColor=GOLD); fill_light=PatternFill("solid",fgColor=LIGHT)
thin=Side(style="thin",color="B0B8C4"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center",vertical="center")
NUM="#,##0.00"; PCT="0.00%"; MULT='0.00"x"'

ER={"Ingresos por venta":(80918236.30,94319491.56),"Costo de ventas":(76876738.28,92204049.57),
 "Utilidad bruta":(4041498.02,2115441.99),"Gastos de operacion":(1894629.56,1984137.34),
 "Utilidad operacional":(2146868.46,131304.65),"Gastos financieros":(2011588.52,1737598.88),
 "Otros ingresos":(252738.42,298668.09),"Utilidad antes 15% Particip.":(388018.36,-1307626.14)}
BG={"Activo corriente":(19201979.20,18676704.08),"Otros activos no corrientes":(3874610.56,1719957.89),
 "Propiedad, planta y equipo (neto)":(14544172.19,14787943.39),"TOTAL ACTIVO":(37620761.95,35184605.36),
 "Pasivo corriente":(14083369.38,9164036.82),"Pasivo no corriente":(13914771.28,17606881.95),
 "TOTAL PASIVO":(27998140.66,26770918.75),"TOTAL PATRIMONIO":(9622621.29,8413686.61)}
ING={"Extraccion":(20354891.91,16710423.95),"Refinacion":(56202818.63,71844163.18),
 "Envasados":(3471005.75,4860256.29),"Suministros/agricolas":(599641.57,606911.16),"Proyecto INIAP":(289878.44,297636.98)}

wb=Workbook()

def title_block(ws, subtitle):
    ws.merge_cells("A1:F1"); c=ws["A1"]; c.value="EXTRACTORA QUEVEPALMA S.A."
    c.font=Font(color="FFFFFF",bold=True,size=14); c.fill=fill_navy; c.alignment=ctr
    ws.merge_cells("A2:F2"); c=ws["A2"]; c.value=subtitle
    c.font=Font(color="FFFFFF",bold=True,size=10); c.fill=fill_steel; c.alignment=ctr
    ws.merge_cells("A3:F3"); c=ws["A3"]; c.value="Quevedo, Los Rios, Ecuador  |  Dolares USD  |  Comparativo 2024 vs 2025"
    c.font=Font(italic=True,size=9,color="555555"); c.alignment=ctr

def hdr_row(ws, r, cols):
    for j,t in enumerate(cols,1):
        c=ws.cell(r,j,t); c.fill=fill_navy; c.font=f_sub; c.alignment=ctr; c.border=border

ws=wb.active; ws.title="Estado de Resultados"
title_block(ws,"ESTADO DE RESULTADOS - Al 31 de diciembre")
hdr_row(ws,5,["Concepto","2024","% 2024","2025","% 2025","Var. %"])
ing24=ER["Ingresos por venta"][0]; ing25=ER["Ingresos por venta"][1]; r=6
for k,(v24,v25) in ER.items():
    bold = k in ("Utilidad bruta","Utilidad operacional","Utilidad antes 15% Particip.")
    ws.cell(r,1,k).font=Font(bold=bold)
    ws.cell(r,2,v24).number_format=NUM
    ws.cell(r,3,v24/ing24).number_format=PCT
    ws.cell(r,4,v25).number_format=NUM
    ws.cell(r,5,v25/ing25).number_format=PCT
    var=(v25/v24-1) if v24 else 0
    e=ws.cell(r,6,var); e.number_format=PCT; e.font=Font(color=GREEN if var>=0 else RED,bold=True)
    if bold:
        for col in range(1,7): ws.cell(r,col).fill=fill_light
        ws.cell(r,2).font=Font(bold=True); ws.cell(r,4).font=Font(bold=True)
    r+=1
ws.freeze_panes="A6"

ws=wb.create_sheet("Balance General")
title_block(ws,"ESTADO DE SITUACION FINANCIERA - Al 31 de diciembre")
hdr_row(ws,5,["Concepto","2024","% Activo","2025","% Activo","Var. %"])
ta24=BG["TOTAL ACTIVO"][0]; ta25=BG["TOTAL ACTIVO"][1]; r=6
for k,(v24,v25) in BG.items():
    tot=k.startswith("TOTAL")
    ws.cell(r,1,k).font=Font(bold=tot)
    ws.cell(r,2,v24).number_format=NUM
    ws.cell(r,3,v24/ta24).number_format=PCT
    ws.cell(r,4,v25).number_format=NUM
    ws.cell(r,5,v25/ta25).number_format=PCT
    var=(v25/v24-1) if v24 else 0
    e=ws.cell(r,6,var); e.number_format=PCT; e.font=Font(color=GREEN if var>=0 else RED,bold=True)
    if tot:
        for col in range(1,7): ws.cell(r,col).fill=fill_gold
        ws.cell(r,2).font=Font(bold=True); ws.cell(r,4).font=Font(bold=True)
    r+=1
ws.freeze_panes="A6"

ws=wb.create_sheet("Ratios Financieros")
title_block(ws,"INDICADORES FINANCIEROS")
def G(d,k,i): return d[k][i]
ratios=[
 ("LIQUIDEZ",None,None,None),
 ("Razon corriente (AC / PC)", G(BG,"Activo corriente",0)/G(BG,"Pasivo corriente",0), G(BG,"Activo corriente",1)/G(BG,"Pasivo corriente",1), MULT),
 ("Capital de trabajo (AC - PC)", G(BG,"Activo corriente",0)-G(BG,"Pasivo corriente",0), G(BG,"Activo corriente",1)-G(BG,"Pasivo corriente",1), NUM),
 ("ENDEUDAMIENTO / SOLVENCIA",None,None,None),
 ("Endeudamiento total (Pasivo / Activo)", G(BG,"TOTAL PASIVO",0)/ta24, G(BG,"TOTAL PASIVO",1)/ta25, PCT),
 ("Apalancamiento (Pasivo / Patrimonio)", G(BG,"TOTAL PASIVO",0)/G(BG,"TOTAL PATRIMONIO",0), G(BG,"TOTAL PASIVO",1)/G(BG,"TOTAL PATRIMONIO",1), MULT),
 ("COBERTURA",None,None,None),
 ("Cobertura de intereses (Ut.Op / G.Financ.)", G(ER,"Utilidad operacional",0)/G(ER,"Gastos financieros",0), G(ER,"Utilidad operacional",1)/G(ER,"Gastos financieros",1), MULT),
 ("RENTABILIDAD",None,None,None),
 ("Margen bruto", G(ER,"Utilidad bruta",0)/ing24, G(ER,"Utilidad bruta",1)/ing25, PCT),
 ("Margen operacional", G(ER,"Utilidad operacional",0)/ing24, G(ER,"Utilidad operacional",1)/ing25, PCT),
 ("Margen neto (antes PT)", G(ER,"Utilidad antes 15% Particip.",0)/ing24, G(ER,"Utilidad antes 15% Particip.",1)/ing25, PCT),
 ("ROA (Utilidad / Activo)", G(ER,"Utilidad antes 15% Particip.",0)/ta24, G(ER,"Utilidad antes 15% Particip.",1)/ta25, PCT),
 ("ROE (Utilidad / Patrimonio)", G(ER,"Utilidad antes 15% Particip.",0)/G(BG,"TOTAL PATRIMONIO",0), G(ER,"Utilidad antes 15% Particip.",1)/G(BG,"TOTAL PATRIMONIO",1), PCT),
]
hdr_row(ws,5,["Indicador","2024","2025","Lectura"]); r=6
for label,v24,v25,fmt in ratios:
    if v24 is None:
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
        c=ws.cell(r,1,label); c.font=f_sub; c.fill=fill_steel
    else:
        ws.cell(r,1,label).font=f_reg
        ws.cell(r,2,v24).number_format=fmt
        b=ws.cell(r,3,v25); b.number_format=fmt
        worse_if_up = ("Endeudamiento" in label or "Apalancamiento" in label)
        better = (v25<v24) if worse_if_up else (v25>v24)
        b.font=Font(color=GREEN if better else RED, bold=True)
        ws.cell(r,4, "mejora" if better else "deterioro").font=Font(italic=True,size=9,color=(GREEN if better else RED))
    r+=1
ws.freeze_panes="A6"

ws=wb.create_sheet("Ventas por Linea")
title_block(ws,"INGRESOS POR LINEA DE PRODUCTO")
hdr_row(ws,5,["Linea","2024","% 2024","2025","% 2025","Var. %"]); r=6
for k,(v24,v25) in ING.items():
    ws.cell(r,1,k)
    ws.cell(r,2,v24).number_format=NUM; ws.cell(r,3,v24/ing24).number_format=PCT
    ws.cell(r,4,v25).number_format=NUM; ws.cell(r,5,v25/ing25).number_format=PCT
    var=v25/v24-1; e=ws.cell(r,6,var); e.number_format=PCT; e.font=Font(color=GREEN if var>=0 else RED,bold=True)
    r+=1
for col in range(1,7): ws.cell(r,col).fill=fill_gold
ws.cell(r,1,"TOTAL INGRESOS").font=Font(bold=True)
ws.cell(r,2,ing24).number_format=NUM; ws.cell(r,2).font=Font(bold=True)
ws.cell(r,3,1).number_format=PCT
ws.cell(r,4,ing25).number_format=NUM; ws.cell(r,4).font=Font(bold=True)
ws.cell(r,5,1).number_format=PCT
ws.cell(r,6,ing25/ing24-1).number_format=PCT; ws.cell(r,6).font=Font(bold=True)
ws.freeze_panes="A6"

for ws in wb.worksheets:
    ws.column_dimensions["A"].width=42
    for col in "BCDEF": ws.column_dimensions[col].width=15
    for row in ws.iter_rows(min_row=5):
        for c in row:
            if c.value is not None and c.row>=5: c.border=border

wb.save(OUT)
print("EXCEL GENERADO:", OUT)
print("Tamano:", os.path.getsize(OUT), "bytes |", len(wb.worksheets), "pestanas")
