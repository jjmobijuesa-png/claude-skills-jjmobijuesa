# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "E:/vars/var 8/3 Analisis Financiero QVP/ANALISIS_FINANCIERO_QUEVEPALMA_2024_2025.xlsx"
NAVY="1B2A4A"; STEEL="4A6FA5"; GOLD="C5A55A"; GREEN="1B7A43"; RED="B02418"; LIGHT="EEF2F7"; YELLOW="FFF2B2"
fill_navy=PatternFill("solid",fgColor=NAVY); fill_steel=PatternFill("solid",fgColor=STEEL)
fill_gold=PatternFill("solid",fgColor=GOLD); fill_light=PatternFill("solid",fgColor=LIGHT); fill_yellow=PatternFill("solid",fgColor=YELLOW)
f_sub=Font(color="FFFFFF",bold=True,size=10)
thin=Side(style="thin",color="B0B8C4"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center",vertical="center"); wrap=Alignment(wrap_text=True,vertical="top")
NUM="#,##0.00"; PCT="0.00%"

def title_block(ws, subtitle):
    ws.merge_cells("A1:E1"); c=ws["A1"]; c.value="EXTRACTORA QUEVEPALMA S.A."
    c.font=Font(color="FFFFFF",bold=True,size=14); c.fill=fill_navy; c.alignment=ctr
    ws.merge_cells("A2:E2"); c=ws["A2"]; c.value=subtitle
    c.font=Font(color="FFFFFF",bold=True,size=10); c.fill=fill_steel; c.alignment=ctr
def hdr(ws,r,cols):
    for j,t in enumerate(cols,1):
        c=ws.cell(r,j,t); c.fill=fill_navy; c.font=f_sub; c.alignment=ctr; c.border=border

wb=load_workbook(OUT)

# ===================== FLUJO DE CAJA 2025 =====================
ws=wb.create_sheet("Flujo de Caja 2025")
title_block(ws,"ESTADO DE FLUJO DE EFECTIVO 2025 (metodo indirecto)")
hdr(ws,4,["Concepto","Monto USD","Nota"])
rows=[
 ("A. ACTIVIDADES DE OPERACION", None, "sec"),
 ("Utilidad neta del ejercicio", -1307626.14, ""),
 ("(+) Depreciacion del ejercicio", 1068958.11, "Δ deprec. acumulada 2024->2025"),
 ("(-) Aumento en capital de trabajo operativo", -703289.96, "AC ex-caja vs PC ex-deuda CP"),
 ("= Flujo de caja operativo (CFO)", -941957.99, "tot"),
 ("B. ACTIVIDADES DE INVERSION", None, "sec"),
 ("(-) Adiciones de propiedad planta y equipo (CapEx)", -1312729.31, "Δ PP&E bruto"),
 ("(+) Disminucion de otros activos no corrientes", 2154652.67, "cobro ctas relacionadas"),
 ("= Flujo de caja de inversion (CFI)", 841923.36, "tot"),
 ("C. ACTIVIDADES DE FINANCIAMIENTO", None, "sec"),
 ("Variacion deuda financiera corto plazo", -3860854.36, "7.13M -> 3.27M"),
 ("Variacion pasivo no corriente (largo plazo)", 3692110.67, "reestructuracion a LP"),
 ("Variacion patrimonio (ex utilidad)", 98691.46, "revaluacion / ajustes"),
 ("= Flujo de caja de financiamiento (CFF)", -70052.23, "tot"),
 ("VARIACION NETA DE EFECTIVO", -170086.86, "fin"),
 ("Efectivo inicial (31-dic-2024)", 521837.55, ""),
 ("Efectivo final (31-dic-2025)", 351750.67, "verifica: cuadra"),
]
r=5
for label,val,tag in rows:
    if tag=="sec":
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
        c=ws.cell(r,1,label); c.font=f_sub; c.fill=fill_steel
    else:
        bold = tag in ("tot","fin")
        ws.cell(r,1,label).font=Font(bold=bold)
        cc=ws.cell(r,2,val); cc.number_format=NUM
        cc.font=Font(bold=bold, color=(GREEN if (val or 0)>=0 else RED))
        ws.cell(r,3,tag if tag not in("tot","fin","")else"").font=Font(italic=True,size=8,color="777777")
        if tag in ("note",): pass
        ws.cell(r,3, [n for n in [label] if False] and "" or (rows[r-5][2] if False else "")).value=None
        ws.cell(r,3,val and "" or "")
        # nota real
        ws.cell(r,3).value = [t for (l,v,t) in [rows[r-5]]][0] if False else None
        ws.cell(r,3).value = tag if tag not in ("tot","fin","sec") else ""
        if tag=="fin":
            for col in range(1,4): ws.cell(r,col).fill=fill_gold
        elif tag=="tot":
            for col in range(1,4): ws.cell(r,col).fill=fill_light
    r+=1
ws.freeze_panes="A5"

# ===================== PROYECCION 2026 (3 escenarios) =====================
ws=wb.create_sheet("Proyeccion 2026")
title_block(ws,"PROYECCION 2026 - 3 ESCENARIOS (supuestos editables en amarillo)")
ing25=94319491.56; opex25=1984137.34
# Supuestos: (crec. ingresos, margen bruto, gastos op, gastos financieros)
sup={
 "Pesimista":(-0.05, 0.020, opex25*1.03, 1750000.0),
 "Base":     ( 0.05, 0.035, opex25*1.03, 1650000.0),
 "Optimista":( 0.12, 0.045, opex25*1.03, 1550000.0),
}
hdr(ws,4,["Concepto","Pesimista","Base","Optimista"])
# fila supuestos editables
def put(r,label,vals,fmt,yellow=False,bold=False,sec=False):
    if sec:
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
        c=ws.cell(r,1,label); c.font=f_sub; c.fill=fill_steel; return
    ws.cell(r,1,label).font=Font(bold=bold)
    for j,v in enumerate(vals,2):
        c=ws.cell(r,j,v); c.number_format=fmt
        if yellow: c.fill=fill_yellow
        if bold: c.font=Font(bold=True)
r=5
put(r,"SUPUESTOS (editables)",None,None,sec=True); r+=1
put(r,"Crecimiento de ingresos %", [sup[s][0] for s in sup], PCT, yellow=True); r+=1
put(r,"Margen bruto %", [sup[s][1] for s in sup], PCT, yellow=True); r+=1
put(r,"Gastos de operacion USD", [sup[s][2] for s in sup], NUM, yellow=True); r+=1
put(r,"Gastos financieros USD", [sup[s][3] for s in sup], NUM, yellow=True); r+=1
put(r,"PROYECCION RESULTANTE",None,None,sec=True); r+=1
ingp=[ing25*(1+sup[s][0]) for s in sup]
put(r,"Ingresos proyectados", ingp, NUM, bold=True); r+=1
ub=[ingp[i]*sup[s][1] for i,s in enumerate(sup)]
put(r,"Utilidad bruta", ub, NUM); r+=1
uo=[ub[i]-sup[s][2] for i,s in enumerate(sup)]
put(r,"Utilidad operacional", uo, NUM, bold=True); r+=1
ut=[uo[i]-sup[s][3] for i,s in enumerate(sup)]
put(r,"Resultado antes de 15% PT", ut, NUM, bold=True)
for j in range(2,5):
    ws.cell(r,j).font=Font(bold=True, color=(GREEN if ut[j-2]>=0 else RED))
r+=1
put(r,"Margen neto %", [ut[i]/ingp[i] for i in range(3)], PCT); r+=1
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r+5,end_column=4)
note=("METODOLOGIA: Ingresos = ventas 2025 x (1 + crecimiento). Utilidad bruta = ingresos x margen bruto supuesto. "
 "Utilidad operacional = bruta - gastos de operacion. Resultado = operacional - gastos financieros. "
 "Los supuestos en AMARILLO son editables. El escenario Base asume recuperacion parcial del margen (2.24% en 2025 hacia 3.5%, "
 "punto medio del bienio) y leve desapalancamiento que reduce el gasto financiero. El escenario Pesimista mantiene el margen comprimido "
 "de 2025; el Optimista se acerca al margen 2024 (4.99%). ALERTA: en TODOS los escenarios la cobertura de intereses sigue siendo el factor "
 "critico - la viabilidad depende de recuperar margen bruto y reducir la carga financiera.")
c=ws.cell(r,1,note); c.alignment=wrap; c.font=Font(size=9,italic=True,color="444444")
ws.freeze_panes="A5"

for ws in wb.worksheets:
    ws.column_dimensions["A"].width=46
    for col in "BCDE": ws.column_dimensions[col].width=16
    for row in ws.iter_rows(min_row=4):
        for c in row:
            if c.value is not None and c.row>=4 and not isinstance(c.value,str) or (isinstance(c.value,str) and c.row==4):
                c.border=border

wb.save(OUT)
print("OK -> Flujo de Caja 2025 + Proyeccion 2026 anadidas")
print("Pestanas:", [s.title for s in wb.worksheets])

# verificacion de cuadre del flujo
tot = -941957.99 + 841923.36 - 70052.23
print(f"CFO+CFI+CFF = {tot:,.2f}  vs  Δcaja real = {351750.67-521837.55:,.2f}  (dif {tot-(351750.67-521837.55):,.4f})")
print("Resultado 2026 antes PT por escenario:")
for i,s in enumerate(sup): print(f"  {s:10s}: {ut[i]:,.0f}")
