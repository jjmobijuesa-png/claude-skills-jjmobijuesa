"""Convierte cada hoja de un .xlsx o .xls a un PDF A4 horizontal independiente.

Uso:
    python convert.py --input book.xlsx --output-dir ./out [--prefix PBC-A5] [--header-tag "Bloque A"]
"""
import argparse
from pathlib import Path

import openpyxl
import xlrd
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER


def safe_name(s):
    return "".join(c if c.isalnum() or c in " -._" else "_" for c in s).strip()


def fmt_val(v):
    if v is None: return ""
    if isinstance(v, float):
        if v == int(v): return str(int(v))
        return f"{v:,.2f}"
    return str(v)


def make_pdf(rows, workbook_stem, sheet_name, header_tag, out_path):
    page_size = landscape(A4)
    margin = 1.2 * cm
    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=page_size,
        leftMargin=margin, rightMargin=margin,
        topMargin=margin, bottomMargin=margin,
        title=workbook_stem,
    )
    page_w = page_size[0] - 2 * margin

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("T", parent=styles["Heading1"],
        alignment=TA_CENTER, fontSize=14, leading=18, spaceAfter=6)
    sub_style = ParagraphStyle("S", parent=styles["Normal"],
        alignment=TA_CENTER, fontSize=9, leading=11,
        textColor=colors.HexColor("#555555"), spaceAfter=6)
    tag_style = ParagraphStyle("Tg", parent=styles["Normal"],
        alignment=TA_CENTER, fontSize=10, leading=12,
        textColor=colors.HexColor("#9c1a1a"), spaceAfter=10)

    story = []
    if header_tag:
        story.append(Paragraph(header_tag, sub_style))
    story.append(Paragraph(workbook_stem, title_style))
    story.append(Paragraph(f"Hoja: <b>{sheet_name}</b>", sub_style))
    story.append(Spacer(1, 4))

    if not rows:
        story.append(Paragraph("(Hoja vacia)", styles["Normal"]))
        doc.build(story)
        return

    max_cols = max(len(r) for r in rows)
    data = [[fmt_val(c) for c in r] + [""] * (max_cols - len(r)) for r in rows]

    ncols = max_cols
    if ncols <= 6: fs = 9
    elif ncols <= 10: fs = 7
    elif ncols <= 14: fs = 6
    else: fs = 5

    col_widths = [page_w / max_cols] * max_cols

    CHUNK = 60
    header = data[0]
    body_rows = data[1:]
    chunks = [body_rows[i:i+CHUNK] for i in range(0, len(body_rows), CHUNK)] or [[]]
    for idx, chunk in enumerate(chunks):
        tbl_data = [header] + chunk
        t = Table(tbl_data, colWidths=col_widths, repeatRows=1, hAlign="CENTER")
        t.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), fs),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#888888")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2a44")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]))
        story.append(t)
        if idx < len(chunks) - 1:
            story.append(PageBreak())
            story.append(Paragraph(f"{workbook_stem} - Hoja {sheet_name} (cont.)", sub_style))

    doc.build(story)


def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        while rows and all(c is None or str(c).strip() == "" for c in rows[-1]):
            rows.pop()
        out[sn] = rows
    return out


def read_xls(path):
    wb = xlrd.open_workbook(str(path))
    out = {}
    for sn in wb.sheet_names():
        sh = wb.sheet_by_name(sn)
        rows = []
        for r in range(sh.nrows):
            rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
        while rows and all(c is None or str(c).strip() == "" for c in rows[-1]):
            rows.pop()
        out[sn] = rows
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--prefix", default="", help="prefix prepended to each output filename")
    ap.add_argument("--header-tag", default="", help="subheading text on each page")
    args = ap.parse_args()

    src = Path(args.input)
    out_dir = Path(args.output_dir); out_dir.mkdir(parents=True, exist_ok=True)
    stem = src.stem
    suffix = src.suffix.lower()

    if suffix == ".xlsx":
        sheets = read_xlsx(src)
    elif suffix == ".xls":
        sheets = read_xls(src)
    else:
        raise ValueError(f"Unsupported extension: {suffix}")

    for sn, rows in sheets.items():
        prefix_part = f"{args.prefix}_" if args.prefix else ""
        out_name = f"{prefix_part}{safe_name(stem)}_{safe_name(sn)}.pdf"
        out_path = out_dir / out_name
        make_pdf(rows, stem, sn, args.header_tag, out_path)
        print(f"OK: {out_path} ({len(rows)} filas)")


if __name__ == "__main__":
    main()
